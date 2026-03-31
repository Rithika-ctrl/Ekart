
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─── STYLES ───────────────────────────────────────────────────────────────────
def hdr(r,g,b): return PatternFill("solid", fgColor=f"{r:02X}{g:02X}{b:02X}")
def bold_font(sz=11, color="000000", bold=True): return Font(bold=bold, size=sz, color=color)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin   = Side(style="thin", color="BBBBBB")
def border(): return Border(left=thin, right=thin, top=thin, bottom=thin)

FILLS = {
    "sheet_tab": hdr(0x1a,0x1a,0x2e),
    "h1":   hdr(0x23,0x37,0x5a),   # dark navy
    "h2":   hdr(0x2e,0x7d,0x32),   # green  – Java
    "h2b":  hdr(0x1a,0x63,0x7e),   # teal   – HTML
    "h3":   hdr(0x01,0x57,0x9b),   # blue
    "cat":  hdr(0x0d,0x47,0xa1),   # deep blue category
    "row_a":hdr(0xf1,0xf8,0xff),   # very light blue  (alternate row A)
    "row_b":hdr(0xff,0xff,0xff),   # white             (alternate row B)
    "yes":  hdr(0xe8,0xf5,0xe9),   # very light green  (working=Yes)
    "no":   hdr(0xff,0xf3,0xe0),   # light orange      (partial/no)
    "warn": hdr(0xff,0xee,0xee),   # pink (missing)
}

WHITE  = Font(bold=True, color="FFFFFF", size=11)
YELLOW = Font(bold=True, color="FFD700", size=10)

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def hrow(ws, row, values, fills, fonts, aligns, heights=None):
    for c, (val, fill, fnt, aln) in enumerate(zip(values, fills, fonts, aligns), 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = fill
        cell.font = fnt
        cell.alignment = aln
        cell.border = border()
    if heights:
        ws.row_dimensions[row].height = heights

def drow(ws, row, values, working=True):
    fill_main = FILLS["yes"] if working else FILLS["no"]
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = fill_main if c == len(values) else (FILLS["row_a"] if row % 2 == 0 else FILLS["row_b"])
        cell.font = Font(size=10)
        cell.alignment = left
        cell.border = border()


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Java Features
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Java Features"
ws1.sheet_properties.tabColor = "1a237e"
ws1.freeze_panes = "A3"

set_col_widths(ws1, [5, 28, 52, 22, 18, 12])

# Title row
ws1.merge_cells("A1:F1")
c = ws1.cell(1, 1, "EKART — Java Backend Feature Inventory (Spring Boot)")
c.fill = FILLS["h1"]; c.font = WHITE; c.alignment = center
ws1.row_dimensions[1].height = 28

# Header row
hdrs = ["#", "Category", "Feature / Endpoint / Class", "File / Class", "Layer", "Working?"]
hrow(ws1, 2, hdrs,
     [FILLS["h3"]]*6,
     [WHITE]*6,
     [center]*6, 22)

rows = []
# ── AUTH & SECURITY ───────────────────────────────────────────────────────────
rows += [
    ("AUTH & SECURITY", "", "", "", ""),
    (1, "Customer Registration", "POST /customer/register — validates email, mobile uniqueness, encrypts password with AES, sends OTP email", "CustomerService / EkartController", "Service+Controller", "Yes"),
    (2, "Customer OTP Email Verification", "Verifies 6-digit OTP, sets verified=true; resends OTP on login if unverified", "CustomerService", "Service", "Yes"),
    (3, "Customer Login", "Session-based login; checks active, verified flags; updates lastLogin timestamp", "CustomerService", "Service", "Yes"),
    (4, "Customer Forgot / Reset Password", "OTP-based reset; validates regex (8+ chars, upper, lower, digit, special)", "CustomerService", "Service", "Yes"),
    (5, "Vendor Registration + OTP", "AES-encrypted password, 6-digit OTP, auto-generates VND-NNNNN code", "VendorService / EkartController", "Service+Controller", "Yes"),
    (6, "Vendor Login / Logout", "Session-based; verifies & active flags; auto-resends OTP if unverified", "VendorService", "Service", "Yes"),
    (7, "Admin Login (env-based)", "Plain-text comparison vs admin.email/admin.password from .env / application.properties", "AdminService", "Service", "Yes"),
    (8, "OAuth2 / Social Login (Google, GitHub, Facebook, Instagram)", "CustomOAuth2UserService loads user; SocialAuthService creates or links account by email/providerId; supports account linking from profile", "CustomOAuth2UserService / SocialAuthService / OAuth2LoginSuccessHandler", "Service+Config", "Yes"),
    (9, "OAuthProviderValidator Middleware", "Role-based provider allowlist: validates which OAuth providers are allowed per role (customer/vendor/admin)", "OAuthProviderValidator", "Config", "Yes"),
    (10, "OAuth2 Account Linking/Unlinking", "GET /customer/link-oauth/{provider}, POST /customer/unlink-oauth", "OAuth2Controller", "Controller", "Yes"),
    (11, "AES Password Encryption / Decryption", "AES/CBC/PKCS5, PBKDF2WithHmacSHA256, 65536 iterations; keys injected from .env via @Value", "AES (helper)", "Helper", "Yes"),
    (12, "JWT Utility (JwtUtil)", "HS256 signed tokens (7-day expiry); generateToken / getCustomerId / getRole / isValid", "JwtUtil (helper)", "Helper", "Yes"),
    (13, "FlutterAuthFilter (JWT for /api/flutter/**)", "OncePerRequestFilter; PUBLIC paths (auth/, products, assistant/chat); admin path guard; X-Customer-Id/X-Vendor-Id/X-Delivery-Id spoofing prevention", "FlutterAuthFilter (middleware)", "Filter", "Yes"),
    (14, "ReactAuthFilter (JWT for /api/react/**)", "Same JWT pattern as FlutterAuthFilter; sets react.userId/react.role attributes", "ReactAuthFilter (middleware)", "Filter", "Yes"),
    (15, "AuthGuard Interceptor (RBAC for Web)", "Protects /admin/*, /api/admin/*, /refund-management/*, /content-management/*, /security-settings/*; checks admin session OR Customer.role==ADMIN", "AuthGuard (middleware)", "Interceptor", "Yes"),
    (16, "DeliveryBoyAuthGuard", "Protects /delivery/** routes; checks verified, active, adminApproved flags in session", "DeliveryBoyAuthGuard (middleware)", "Interceptor", "Yes"),
    (17, "IndiaOnlyInterceptor (Geo-blocking)", "ip-api.com lookup; in-memory IP cache; CORS-safe (OPTIONS exempt); private IPs always allowed; fail-open on error; JSON for AJAX, redirect for browser", "IndiaOnlyInterceptor (middleware)", "Interceptor", "Yes"),
    (18, "SecurityConfig (Spring Security)", "Multi-chain: Chain 1 = /api/flutter/** + /api/react/** (stateless JWT); Chain 2 = Thymeleaf pages (session); OAuth2 login integration", "SecurityConfig (config)", "Config", "Yes"),
    (19, "Spring Security Session Management", "Session invalidate on logout; session attrs: customer, vendor, admin, deliveryBoy, guest", "CustomerService / VendorService / AdminService / DeliveryBoyService", "Service", "Yes"),
    (20, "Account Active/Inactive Toggle (Admin)", "Admin suspends/activates customer accounts; suspended users see failure on login", "AdminService / AdminAccountService", "Service", "Yes"),
    (21, "RBAC Roles Enum", "Three roles: CUSTOMER, ORDER_MANAGER, ADMIN; stored as @Enumerated(STRING) on Customer entity", "Role (dto) / Customer (dto)", "DTO", "Yes"),
    (22, "Delivery Boy Self-Registration + OTP + Admin Approval Workflow", "Self-register → email OTP → admin notified → admin approves/rejects; admin-created accounts skip approval; login flow checks all states", "DeliveryBoyService / DeliveryAdminService", "Service", "Yes"),
    ("AUTH & SECURITY — REACT/FLUTTER API AUTH", "", "", "", ""),
    (23, "Flutter API Auth — Customer/Vendor/Admin/Delivery Login", "POST /api/flutter/auth/{role}/login — AES decryption, Base64 token (flutter) or JWT (react), all state checks", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (24, "Flutter API Registration", "POST /api/flutter/auth/customer/register, /auth/vendor/register, /auth/delivery/register", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (25, "React API Forgot Password / OTP / Reset (3-step)", "Step1: send OTP; Step2: verify OTP → in-memory otpVerified map; Step3: reset password & consume flag (one-use)", "ReactApiController", "Controller", "Yes"),
    (26, "Delivery Boy Statuses on Login", "active / pending / unverified / inactive — distinct JSON status field with appropriate HTTP codes", "FlutterApiController / ReactApiController", "Controller", "Yes"),
]

rows += [
    ("PRODUCT MANAGEMENT", "", "", "", ""),
    (27, "Product Entity (full)", "Fields: name, description, price, mrp, category, stock, stockAlertThreshold, imageLink, extraImageLinks, videoLink, approved, allowedPinCodes, vendor (ManyToOne), reviews (OneToMany)", "Product (dto)", "DTO", "Yes"),
    (28, "Vendor Add Product (Web)", "Cloudinary image + multiple extra images + video upload; auto-unapproved; stock alert check", "VendorService / EkartController", "Service+Controller", "Yes"),
    (29, "Vendor Edit / Update Product", "Edit all fields; new Cloudinary upload optional; stock delta triggers BackInStock notification", "VendorService", "Service", "Yes"),
    (30, "Vendor Delete Product", "Deletes linked Items in cart first, then product", "VendorService", "Service", "Yes"),
    (31, "Admin Approve / Reject Product (Web)", "Toggle product.approved; endpoint /admin/approve-products/{id}", "AdminService / EkartController", "Service+Controller", "Yes"),
    (32, "Product Approval (API — Flutter/React)", "POST /api/flutter/admin/products/{id}/approve|reject + bulk approve-all", "FlutterApiController / FlutterAdminApiController", "Controller", "Yes"),
    (33, "MRP / Discount Display", "getDiscountPercent() = (mrp-price)/mrp*100; isDiscounted() helper; shown on product cards", "Product (dto)", "DTO", "Yes"),
    (34, "Extra Images (Gallery)", "Comma-separated URLs in extraImageLinks; getExtraImageList() splits to List<String>", "Product (dto)", "DTO", "Yes"),
    (35, "Product Video Upload", "Cloudinary video upload with resource_type=video; videoLink stored on Product", "VendorService / CloudinaryHelper", "Service+Helper", "Yes"),
    (36, "Pin Code Delivery Restriction (Product)", "allowedPinCodes comma-separated; isDeliverableTo(pin) checks; filterValidPins() validates against Indian postal circle rules", "Product (dto) / PinCodeValidator (helper)", "DTO+Helper", "Yes"),
    (37, "PinCodeValidator", "Indian 6-digit PIN validation: must be exactly 6 digits AND start with a valid postal circle prefix (11-19, 20-28, 30-34, 36-39, 40-49, 50-53, 56-59, 60-66, 67-69, 70-74, 75-77, 78-79, 80-85, 90-99)", "PinCodeValidator (helper)", "Helper", "Yes"),
    (38, "Product Search (multi-field)", "Name + description + category containing (case-insensitive); HashSet dedup; fuzzy correction fallback", "CustomerService / ProductRepository", "Service+Repo", "Yes"),
    (39, "Fuzzy Search / Spell Correction", "Levenshtein distance; adaptive threshold by query length; per-word matching for multi-word names; SharedPrefixLength tie-breaking", "SearchService", "Service", "Yes"),
    (40, "Search Suggestions (AJAX)", "GET /api/search/suggestions?q=; returns up to 8 suggestions with name, category, imageLink, purchaseCount", "SearchController / SearchService", "Controller+Service", "Yes"),
    (41, "Search Fuzzy Endpoint", "GET /api/search/fuzzy?q=; returns best corrected term or empty string", "SearchController / SearchService", "Controller+Service", "Yes"),
    (42, "Product CSV Upload (Vendor API)", "POST /api/react/vendor/products/upload-csv; CSV parsing with quote handling; upsert by id or create new; up to 50 errors reported", "ReactApiController", "Controller", "Yes"),
    (43, "Vendor Product CRUD API (Flutter/React)", "Add, Update, Delete, List products via REST; vendor ownership enforced", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (44, "Product Categories (dynamic)", "Derived from all approved products; distinct+sorted; GET /api/flutter/products/categories", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (45, "Category Entity (hierarchical)", "Parent (top-level) → Sub-category; emoji + displayOrder; products link by sub-category name string", "Category (dto) / CategoryService", "DTO+Service", "Yes"),
    (46, "Product Detail View (Similar Products)", "Loads 2 approved products from same category as related products on detail page", "CustomerService", "Service", "Yes"),
    (47, "Product Filters (React API)", "GET /api/react/products?search=&category=&minPrice=&maxPrice=&sortBy=; server-side filtering + sorting (price_asc, price_desc, name)", "ReactApiController", "Controller", "Yes"),
    (48, "Product Average Rating (computed)", "getAverageRating() rounds to nearest 0.5 across all reviews; getReviewCount() safe null", "Product (dto)", "DTO", "Yes"),
]

rows += [
    ("CART & CHECKOUT", "", "", "", ""),
    (49, "Cart Entity", "OneToOne with Customer (CascadeType.ALL); OneToMany Items (EAGER, orphanRemoval=true); Serializable", "Cart (dto)", "DTO", "Yes"),
    (50, "Add to Cart (Web)", "Checks stock, prevents duplicate by name, decrements stock, sets unitPrice snapshot", "CustomerService", "Service", "Yes"),
    (51, "Cart AJAX Quantity Increase/Decrease/Remove", "Returns updated quantity, lineTotal, cartTotal, freeDelivery flag, deliveryCharge in JSON", "CustomerService (ajaxIncrease/ajaxDecrease/ajaxRemove)", "Service", "Yes"),
    (52, "Cart View Page", "Calculates lineTotal per item (unitPrice × qty); totals with delivery charge", "CustomerService", "Service", "Yes"),
    (53, "Stock Restoration on Remove", "When item removed from cart, productRepository restores stock += quantity", "CustomerService", "Service", "Yes"),
    (54, "Cart API (Flutter/React)", "GET /cart, POST /cart/add, DELETE /cart/remove/{id}, PUT /cart/update; returns subtotal, couponDiscount, deliveryCharge, total", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (55, "Coupon System", "Coupon entity: code (unique), type (PERCENT/FLAT), value, minOrderAmount, maxDiscount, usageLimit, usedCount, expiryDate, active; isValid(), calculateDiscount()", "Coupon (dto) / CouponRepository", "DTO+Repo", "Yes"),
    (56, "Coupon Apply (React API)", "POST /api/react/cart/coupon; validates, checks minOrderAmount, calculates discount; stores in ConcurrentHashMap per customerId", "ReactApiController", "Controller", "Yes"),
    (57, "Coupon Remove (React API)", "DELETE /api/react/cart/coupon; clears in-memory coupon for customer", "ReactApiController", "Controller", "Yes"),
    (58, "Active Coupons List (React API)", "GET /api/react/coupons — returns only valid (active, not expired, within usageLimit) coupons", "ReactApiController", "Controller", "Yes"),
    (59, "Coupon Admin CRUD (React API)", "GET/POST/toggle/DELETE /api/react/admin/coupons/** — full CRUD with code uniqueness check", "ReactApiController", "Controller", "Yes"),
    (60, "Coupon Admin CRUD (Web)", "EkartController /admin/coupons — web form CRUD", "EkartController / CouponController", "Controller", "Yes"),
    (61, "Payment Page — GST Breakdown", "GstUtil.calculateTotalGst() across all cart items; getMixedGstLabel() for display; passed to payment.html template", "CustomerService / GstUtil", "Service+Helper", "Yes"),
    (62, "GstUtil — Indian GST Rate Engine", "60+ category keyword→rate mappings: 0%, 5%, 12%, 18%; back-calculates GST from inclusive price: GST = price × rate / (1+rate)", "GstUtil (helper)", "Helper", "Yes"),
    (63, "Delivery Charge Logic", "Free (₹0) for orders ≥ ₹500; ₹40 for < ₹500; applied consistently in web, flutter, react", "CustomerService / ReactApiController / FlutterApiController", "Service+Controller", "Yes"),
    (64, "Order Placement — Multi-Vendor Split", "Cart grouped by vendor → one sub-order per vendor; all share parentOrderId; delivery fee on first sub-order only; stock decremented", "CustomerService.paymentSuccess()", "Service", "Yes"),
    (65, "COD + Razorpay Payment Mode", "paymentMode stored on Order; Razorpay key injected from @Value; payment.html has both flows", "CustomerService / EkartController", "Service+Controller", "Yes"),
    (66, "PIN Code Validation at Checkout", "Checks product.isDeliverableTo(pin) for each cart item; blocks order if restricted product can't deliver to entered PIN", "CustomerService.paymentSuccess()", "Service", "Yes"),
    (67, "Delivery Address Snapshot", "Builds formatted string from customer's last address at checkout; stored on Order.deliveryAddress", "CustomerService.paymentSuccess()", "Service", "Yes"),
    (68, "Order Confirmation Email", "Sends HTML email with orderId, amount, paymentMode, deliverySlot, items list (EmailSender.sendOrderConfirmation)", "CustomerService / EmailSender", "Service+Helper", "Yes"),
    (69, "Cart API — Coupon + Delivery Integration (React)", "GET /api/react/cart returns subtotal, couponDiscount, deliveryCharge (0 if ≥₹500 post-discount, else 40), total", "ReactApiController", "Controller", "Yes"),
]

rows += [
    ("ORDER MANAGEMENT", "", "", "", ""),
    (70, "Order Entity", "Fields: amount, dateTime, deliveryCharge, totalPrice, gstAmount, paymentMode, orderDate (@CreationTimestamp), deliveryTime, replacementRequested, trackingStatus, currentCity, items (CascadeType.ALL EAGER), customer, warehouse (LazyL), deliveryBoy (Lazy), deliveryPinCode, deliveryAddress, parentOrderId, vendorId, vendorName", "Order (dto)", "DTO", "Yes"),
    (71, "TrackingStatus Enum", "PROCESSING → PACKED → SHIPPED → OUT_FOR_DELIVERY → DELIVERED → REFUNDED → CANCELLED; getProgressPercent() for UI progress bar; getStepIndex() for ordering", "TrackingStatus (dto)", "DTO", "Yes"),
    (72, "View Orders (Web)", "Loads all customer orders; returnEligibleMap (7-day window); replacementRequestedMap", "CustomerService.viewOrders()", "Service", "Yes"),
    (73, "Order History Page", "Separate order-history.html page; same data as view-orders", "CustomerService.viewOrderHistory()", "Service", "Yes"),
    (74, "Cancel Order (Web)", "Restores product stock by name search; sends cancellation email; deletes order", "CustomerService.cancelOrder()", "Service", "Yes"),
    (75, "Request Replacement (Web)", "7-day eligibility window; sets replacementRequested=true; sends replacement email", "CustomerService.requestReplacement()", "Service", "Yes"),
    (76, "Order Cancel API (Flutter/React)", "POST /api/flutter/orders/{id}/cancel + /api/react/orders/{id}/cancel; restores stock; validates delivery/cancelled status", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (77, "Reorder (API)", "POST /api/react/orders/{id}/reorder + /api/flutter/orders/{id}/reorder; clears cart, adds in-stock items at current prices with capped qty", "ReactApiController / FlutterApiController / ReorderService", "Controller+Service", "Yes"),
    (78, "Reorder — Check Stock Pre-flight", "GET /api/orders/{id}/check-stock — returns per-item availability status without modifying cart", "ReorderApiController / ReorderService", "Controller+Service", "Yes"),
    (79, "Sub-order Grouping (parentOrderId)", "Multi-vendor carts create N sub-orders sharing parentOrderId; OrderRepository.findByParentOrderId()", "Order (dto) / OrderRepository", "DTO+Repo", "Yes"),
    (80, "Vendor Views Orders (Web)", "Loads orders containing vendor's products; split into pending (PROCESSING), inProgress, delivered", "VendorService.loadVendorOrders()", "Service", "Yes"),
    (81, "Vendor Mark Order as Packed (Web + API)", "POST /vendor/order/{id}/ready → PROCESSING→PACKED; logs TrackingEventLog; Web + /api/react/vendor/orders/{id}/mark-packed", "VendorService.markOrderReady() / ReactApiController", "Service+Controller", "Yes"),
    (82, "Vendor Orders API (Flutter/React)", "GET /api/flutter/vendor/orders — returns vendor-specific items per order + vendorTotal", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (83, "Admin Update Order Status (API)", "POST /api/flutter/admin/orders/{id}/status body:{status} — any TrackingStatus enum value", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (84, "Admin Delivery Management Page", "Lists PACKED orders (ready for assignment), SHIPPED, OUT_FOR_DELIVERY; shows pending delivery boy approvals", "DeliveryAdminService", "Service", "Yes"),
    (85, "Admin Assign Delivery Boy to Order", "POST /admin/delivery/assign — validates PACKED status, approves+active boy; transitions to SHIPPED; sends shipped email; logs event", "DeliveryAdminService / FlutterAdminApiController", "Service+Controller", "Yes"),
    (86, "Eligible Delivery Boys for Order (AJAX)", "GET /admin/delivery/boys/for-order/{orderId} — filters by order pin code (exact pin → warehouse fallback → all active)", "DeliveryAdminService", "Service", "Yes"),
    (87, "Report Issue / Dispute", "POST /api/react/orders/{id}/report-issue — structured audit log to stdout + HTML email to admin; always returns 200", "ReactApiController", "Controller", "Yes"),
    (88, "Order Tracking Page (Web)", "GET /api/orders/{id}/track — real event-based history from TrackingEventLog; progressPercent from TrackingStatus", "OrderTrackingController / OrderTrackingService", "Controller+Service", "Yes"),
    (89, "Order Tracking API (React JWT)", "GET /api/react/orders/{id}/track — JWT auth; returns history list, progressPercent, estimatedDelivery (orderDate+48h)", "ReactApiController", "Controller", "Yes"),
    (90, "TrackingEventLog Entity", "One row per real status change; order (ManyToOne), status, city, description, updatedBy, eventTime; inserted by system/vendor/admin/delivery_boy", "TrackingEventLog (dto) / TrackingEventLogRepository", "DTO+Repo", "Yes"),
]

rows += [
    ("DELIVERY SYSTEM", "", "", "", ""),
    (91, "DeliveryBoy Entity", "Fields: name, email (unique index), mobile, password (AES), otp, verified, adminApproved, active, warehouse (Lazy ManyToOne), assignedPinCodes (CSV), deliveryBoyCode (unique); covers(pin) helper", "DeliveryBoy (dto)", "DTO", "Yes"),
    (92, "Warehouse Entity", "Fields: name, city, state, warehouseCode (auto WH-NNN), servedPinCodes (CSV up to 5000 chars), active; serves(pin) helper", "Warehouse (dto)", "DTO", "Yes"),
    (93, "Warehouse Auto-Assignment on Order", "Finds warehouse whose servedPinCodes covers customer's deliveryPinCode; delimiter-aware SQL LIKE matching", "CustomerService.paymentSuccess() / WarehouseRepository", "Service+Repo", "Yes"),
    (94, "Delivery Boy Dashboard (Web)", "Shows toPickUp (SHIPPED), outNow (OUT_FOR_DELIVERY), delivered; warehouse info; pending change request status", "DeliveryBoyService.loadHome()", "Service", "Yes"),
    (95, "Delivery Boy Pickup (SHIPPED→OUT_FOR_DELIVERY)", "POST /delivery/order/{id}/pickup — validates assignment, generates 6-digit OTP, sends email to customer, logs event", "DeliveryBoyService.markPickedUp()", "Service", "Yes"),
    (96, "Delivery OTP Confirmation", "POST /delivery/order/{id}/confirm — verifies 6-digit OTP at doorstep, marks DELIVERED, logs event, sends delivery confirmation email", "DeliveryBoyService.confirmDelivery()", "Service", "Yes"),
    (97, "DeliveryOtp Entity", "OneToOne with Order (unique constraint); otp (int), used, generatedAt, usedAt", "DeliveryOtp (dto) / DeliveryOtpRepository", "DTO+Repo", "Yes"),
    (98, "Delivery OTP Email (to Customer)", "HTML email via EmailSender.sendDeliveryOtp(); sent when delivery boy picks up order", "EmailSender", "Helper", "Yes"),
    (99, "Warehouse Change Request", "DeliveryBoy submits request; one pending at a time; admin approves/rejects; changes warehouse + clears assignedPinCodes; approval/rejection emails sent", "WarehouseChangeRequest (dto) / DeliveryBoyService", "DTO+Service", "Yes"),
    (100, "Admin Warehouse Management Page", "GET/POST /admin/delivery/warehouse — CRUD with warehouseCode auto-generation", "DeliveryAdminService / EkartController", "Service+Controller", "Yes"),
    (101, "Delivery Boy API (Flutter/React)", "Login, pickup, deliver, warehouse change request, warehouse list — all stateless with X-Delivery-Boy-Id or X-Delivery-Id header", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (102, "Delivery Home API (Flutter)", "GET /api/flutter/delivery/home — profile + toPickUp + outNow + delivered; hasPendingWarehouseRequest flag", "FlutterApiController", "Controller", "Yes"),
    (103, "Admin Delivery Boy Approve/Reject API", "POST /api/flutter/admin/delivery/boy/approve|reject — stateless JWT-auth equivalent of web session endpoint", "FlutterAdminApiController", "Controller", "Yes"),
    (104, "Admin Delivery Assign API (Flutter)", "POST /api/flutter/admin/delivery/assign — PACKED→SHIPPED; validates all delivery boy flags", "FlutterAdminApiController", "Controller", "Yes"),
]

rows += [
    ("REFUND SYSTEM", "", "", "", ""),
    (105, "Refund Entity", "Fields: order (ManyToOne), customer (ManyToOne), amount, reason, status (PENDING/APPROVED/REJECTED), rejectionReason, requestedAt (@CreationTimestamp), processedAt, processedBy", "Refund (dto) / RefundRepository", "DTO+Repo", "Yes"),
    (106, "RefundImage Entity", "Stores Cloudinary URLs for evidence images (max 5 per refund); ManyToOne→Refund", "RefundImage (dto) / RefundImageRepository", "DTO+Repo", "Yes"),
    (107, "Admin Refund Management (Web)", "Approve (REFUNDED status on Order, clears replacementRequested) / Reject (with mandatory reason); sends notification", "RefundService / AdminRefundController", "Service+Controller", "Yes"),
    (108, "Refund Request API (Flutter/React)", "POST /api/flutter/refund/request — type (REFUND/REPLACEMENT) prepended to reason; validates DELIVERED status", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (109, "Refund Status API", "GET /api/flutter/refund/status/{orderId} — parses [REFUND]/[REPLACEMENT] prefix from stored reason", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (110, "Refund Evidence Image Upload (API)", "POST /api/react/refund/{id}/upload-image — multipart, max 5 images, JPEG/PNG/WEBP only, max 5MB each, countByRefundId cap check", "ReactApiController", "Controller", "Yes"),
    (111, "Admin Refund Approve/Reject (React API)", "POST /api/react/admin/refunds/{id}/approve|reject — delegates to RefundService; updates Order.trackingStatus=REFUNDED", "ReactApiController / RefundService", "Controller+Service", "Yes"),
    (112, "Legacy replacementRequested Migration", "RefundService.migrateExistingReplacementRequests() converts old replacementRequested=true orders to Refund entities", "RefundService", "Service", "Yes"),
]

rows += [
    ("REVIEWS", "", "", "", ""),
    (113, "Review Entity", "Fields: rating (1-5), comment (1000 chars), customerName, createdAt (@PrePersist), product (ManyToOne); getStarDisplay() helper", "Review (dto)", "DTO", "Yes"),
    (114, "ReviewImage Entity", "Cloudinary URL evidence images for reviews (max 5); ManyToOne→Review", "ReviewImage (dto) / ReviewImageRepository", "DTO+Repo", "Yes"),
    (115, "Add Review (Web)", "Duplicate check by customerName+productId; rating clamped 1-5", "CustomerService.addReview()", "Service", "Yes"),
    (116, "Add Review API (Flutter/React)", "POST /api/flutter/reviews/add — customer name from authenticated user", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (117, "Review Image Upload API", "POST /api/react/reviews/{id}/upload-image — multipart, max 5, JPEG/PNG/WEBP, 5MB cap; countByReviewId check", "ReactApiController", "Controller", "Yes"),
    (118, "Product Reviews Endpoint", "GET /api/flutter/products/{id}/reviews — avgRating, reviewCount, list", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (119, "Admin Review Management (Web)", "View, delete, bulk-delete by productName; star distribution (1-5 counts)", "AdminReviewController", "Controller", "Yes"),
    (120, "Admin Review Management (API)", "GET /api/flutter/admin/reviews?filter=&search=; DELETE by id; POST bulk-delete by productName", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (121, "Overall Average Rating (Repo)", "ReviewRepository.getOverallAverageRating() — JPQL AVG query", "ReviewRepository", "Repository", "Yes"),
]

rows += [
    ("WISHLIST", "", "", "", ""),
    (122, "Wishlist Entity", "UniqueConstraint(customer_id, product_id); addedAt timestamp", "Wishlist (dto) / WishlistRepository", "DTO+Repo", "Yes"),
    (123, "Wishlist Toggle (Web AJAX)", "POST /api/wishlist/toggle — adds or removes; returns added bool, wishlistCount", "WishlistController / WishlistService", "Controller+Service", "Yes"),
    (124, "Wishlist Page (Web)", "GET /account/wishlist — all wishlist products with count", "WishlistController", "Controller", "Yes"),
    (125, "Wishlist Remove (Web)", "DELETE /api/wishlist/{productId}", "WishlistController", "Controller", "Yes"),
    (126, "Wishlist IDs Endpoint (Web)", "GET /api/wishlist/ids — Set<Integer> productIds for checked heart icons", "WishlistController", "Controller", "Yes"),
    (127, "Wishlist API (Flutter/React)", "GET /wishlist, POST /wishlist/toggle, GET /wishlist/ids — all customer-auth protected", "FlutterApiController / ReactApiController", "Controller", "Yes"),
]

rows += [
    ("BACK-IN-STOCK NOTIFICATIONS", "", "", "", ""),
    (128, "BackInStockSubscription Entity", "UniqueConstraint(customer_id, product_id); subscribedAt, notified bool, notifiedAt; prevents duplicate subscriptions", "BackInStockSubscription (dto)", "DTO", "Yes"),
    (129, "Subscribe / Unsubscribe (API)", "POST /notify-me/{id}, DELETE /notify-me/{id}, GET /notify-me/{id} — resolves customer from session, JWT, or X-Customer-Id header", "BackInStockController / BackInStockService", "Controller+Service", "Yes"),
    (130, "Notify All Subscribers (Auto-trigger)", "Called in VendorService.updateProduct() when oldStock==0 and newStock>0; sends HTML email to each unnotified subscriber; marks notified=true", "BackInStockService.notifySubscribers()", "Service", "Yes"),
    (131, "Back-in-Stock Email", "EmailSender.sendBackInStockNotification() — HTML email with product name, image, price, stock, direct link", "EmailSender", "Helper", "Yes"),
    (132, "Subscriber Count (for Vendor/Admin)", "BackInStockRepository.countPendingByProduct() — JPQL COUNT query", "BackInStockRepository", "Repository", "Yes"),
    (133, "Flutter Notify-Me API", "POST/DELETE/GET /api/flutter/notify-me/{productId} — identical logic via FlutterApiController", "FlutterApiController", "Controller", "Yes"),
]

rows += [
    ("STOCK ALERTS", "", "", "", ""),
    (134, "StockAlert Entity", "product, vendor, stockLevel, alertTime, emailSent, acknowledged, message", "StockAlert (dto) / StockAlertRepository", "DTO+Repo", "Yes"),
    (135, "Auto Stock Alert Check", "checkStockLevel() called on every product save/update; creates alert if stock ≤ threshold and no unacknowledged alert exists", "StockAlertService", "Service", "Yes"),
    (136, "Stock Alert Email to Vendor", "EmailSender.sendStockAlert() — HTML email with productName, currentStock, threshold, productId", "StockAlertService / EmailSender", "Service+Helper", "Yes"),
    (137, "Vendor View Stock Alerts (Web)", "GET /stock-alerts — sorted unacknowledged first, then by id desc; count badge", "StockAlertService / EkartController", "Service+Controller", "Yes"),
    (138, "Acknowledge Alert", "POST /stock-alerts/acknowledge/{id}; validates vendor ownership", "StockAlertService / EkartController", "Service+Controller", "Yes"),
    (139, "Stock Alerts API (Flutter/React)", "GET /api/flutter/vendor/stock-alerts — sorted unacknowledged first; POST /{id}/acknowledge; unacknowledgedCount field", "FlutterApiController / ReactApiController", "Controller", "Yes"),
]

rows += [
    ("PROFILE & ADDRESS", "", "", "", ""),
    (140, "Customer Profile Entity", "Fields: name, email, mobile, password, otp, verified, role, active, lastLogin, provider, providerId, cart (OneToOne CascadeType.ALL), addresses (OneToMany), recentlyViewedProducts (CSV), profileImage (Cloudinary URL)", "Customer (dto)", "DTO", "Yes"),
    (141, "Address Entity (Structured)", "Fields: recipientName, houseStreet, city, state, postalCode, details (legacy fallback); getFormattedAddress() builds multi-line string; ManyToOne→Customer", "Address (dto)", "DTO", "Yes"),
    (142, "Address Page / CRUD (Web)", "GET /customer/address — list saved addresses; POST save structured address; DELETE remove by id", "CustomerService / EkartController", "Service+Controller", "Yes"),
    (143, "Profile API — Get/Update/Change Password", "GET /api/flutter/profile, PUT /api/flutter/profile/update, PUT /api/flutter/profile/change-password — customer + vendor variants", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (144, "Profile Address Add/Delete (API)", "POST /api/flutter/profile/address/add — structured or legacy flat; PIN validation; DELETE /profile/address/{id}/delete", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (145, "Profile Image Upload (Web)", "POST /customer/profile/upload-image — Cloudinary upload, session customer updated", "CustomerImageUploadController", "Controller", "Yes"),
    (146, "Profile Image Upload (API)", "POST /api/profile/upload-image — JWT auth; Cloudinary; customer profileImage updated; GET /api/profile/remove-image", "ProfileWishlistApiController", "Controller", "Yes"),
    (147, "Delete Account (Customer Web)", "Deletes wishlist, refunds, orders, then customer; session invalidated", "CustomerService.deleteAccount()", "Service", "Yes"),
    (148, "Vendor Profile API", "GET /api/flutter/vendor/profile, PUT /api/flutter/vendor/profile/update, PUT /api/flutter/vendor/profile/change-password", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (149, "Vendor Storefront Update API", "PUT /api/react/vendor/storefront/update — name, mobile (with uniqueness check), description; all fields optional", "ReactApiController", "Controller", "Yes"),
    (150, "Recently Viewed Products", "POST /api/recently-viewed/sync — stores comma-separated productIds in Customer.recentlyViewedProducts; GET /api/recently-viewed/sync — loads back; GET /api/recently-viewed/products?ids= — resolves product summaries", "RecentlyViewedController", "Controller", "Yes"),
]

rows += [
    ("BANNERS & CONTENT", "", "", "", ""),
    (151, "Banner Entity", "Fields: title, imageUrl, linkUrl, active, showOnHome (landing page), showOnCustomerHome (post-login), displayOrder", "Banner (dto) / BannerRepository", "DTO+Repo", "Yes"),
    (152, "Banner CRUD (Admin Web)", "Add with Cloudinary upload OR URL; toggle active, showOnHome, showOnCustomerHome; delete; update", "BannerService / EkartController", "Service+Controller", "Yes"),
    (153, "Banner Display — Landing Page", "findByActiveTrueAndShowOnHomeTrueOrderByDisplayOrderAsc()", "BannerService / BannerRepository", "Service+Repo", "Yes"),
    (154, "Banner Display — Customer Home", "@Cacheable('banners-home'); findByActiveTrueAndShowOnCustomerHomeTrueOrderByDisplayOrderAsc()", "BannerService", "Service", "Yes"),
    (155, "Banners API (Public)", "GET /api/flutter/banners (customer home), GET /api/flutter/home-banners (landing page)", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (156, "Admin Banner CRUD API", "GET/POST-add/toggle/toggle-customer-home/DELETE + toggle-home + update — full admin banner management", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (157, "Policy Entity + CRUD", "Model: title, content, category, slug (unique), lastUpdated, authorAdminId; CRUD via REST /api/policies; slug-based lookup", "Policy (model) / PolicyController / PolicyRepository", "Controller+Model+Repo", "Yes"),
    (158, "Policy Audit Log", "AuditLogService.logPolicyAction() — SLF4J INFO logging for CREATED/UPDATED/DELETED", "AuditLogService", "Service", "Yes"),
]

rows += [
    ("ANALYTICS & REPORTING", "", "", "", ""),
    (159, "Dual Database Architecture", "Main DB: MySQL (Railway) — all transactional entities; Reporting DB: H2 file — SalesRecord only; separate EntityManagerFactory beans; MainDataSourceConfig + ReportingDataSourceConfig", "MainDataSourceConfig / ReportingDataSourceConfig (reporting)", "Config", "Yes"),
    (160, "SalesRecord Entity (Reporting DB)", "One row per item sold: orderId, orderDate, productId, productName, category, itemPrice, quantity, vendorId, vendorName, customerId, customerName, orderTotal, deliveryCharge", "SalesRecord (reporting)", "Reporting", "Yes"),
    (161, "ReportingService.recordOrder()", "Called after transaction commit via TransactionSynchronizationManager; idempotent by orderId; resolves vendor via main DB productRepository", "ReportingService", "Service", "Yes"),
    (162, "Vendor Sales Report (Web)", "Daily/Weekly/Monthly/Overall summaries from reporting DB; top products, category revenue; saves SalesReport snapshots; JSON for chart rendering", "VendorService.loadSalesReport() / ReportingService", "Service", "Yes"),
    (163, "Reporting DB Sync Endpoint", "POST /vendor/sync-reporting-db — backfills reporting DB from all vendor orders in main DB", "VendorService.syncReportingDb()", "Service", "Yes"),
    (164, "Admin Platform Analytics (Web)", "GET /admin/analytics — platform-wide: customers, vendors, products (pending/approved), orders (by status), revenue, daily orders (7-day), category stats", "AdminService.analytics()", "Service", "Yes"),
    (165, "Admin Analytics API (React)", "GET /api/react/admin/analytics — richer data: dailyOrders (7-day map), monthlyRevenue (6-month map), top 5 products by revenue (with unitsSold), categoryStats, statusBreakdown", "ReactApiController", "Controller", "Yes"),
    (166, "Vendor Sales Report API (React)", "GET /api/react/vendor/sales-report?period=daily|weekly|monthly|yearly — bucket data (labels+revenue), topProducts by units, pendingOrders count, avgOrderValue", "ReactApiController", "Controller", "Yes"),
    (167, "Admin User Spending Analytics", "GET /api/react/admin/spending — per-customer: totalSpent, totalOrders, avgOrderValue, topCategory, categorySpending{}, monthlySpending{}; sorted by totalSpent desc", "ReactApiController", "Controller", "Yes"),
    (168, "Customer Spending Analytics (Web)", "GET /api/account/spending-summary — totalSpent, totalOrders, avgOrderValue, topCategory, monthlySpending, categorySpending (DELIVERED orders only)", "SpendingAnalyticsController / SpendingAnalyticsService", "Controller+Service", "Yes"),
    (169, "Customer Spending API (Flutter/React)", "GET /api/flutter/spending-summary — same logic; monthlySpending per YYYY-MM key", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (170, "Vendor Stats API", "GET /api/flutter/vendor/stats — totalRevenue (non-cancelled), totalOrders, totalProducts, activeProducts, lowStockProducts", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (171, "Admin Platform Stats API (Flutter)", "GET /api/flutter/admin/stats — totalCustomers, totalVendors, totalProducts, pendingProducts, totalOrders, pendingOrders, totalRevenue (DELIVERED), deliveryBoys, pendingApprovals, pendingWHChanges, reviews, banners", "FlutterApiController", "Controller", "Yes"),
]

rows += [
    ("AI ASSISTANT", "", "", "", ""),
    (172, "AI Assistant — Gemini API Integration", "Gemini 1.5 Flash via HTTP; system prompt with live contextBlock; conversation history; temperature=0.3; 12s timeout; fallback to local", "AiAssistantService", "Service", "Yes"),
    (173, "Role-specific System Prompts", "CRITICAL RULES injected: use LIVE DATA, role isolation (vendor can't answer customer questions), Ekart facts (delivery free ≥₹500, order stages, return window)", "AiAssistantService.buildSystemPrompt()", "Service", "Yes"),
    (174, "Context Block Builder (Customer)", "Builds: customer name, cart (items, lineTotals, delivery hint), orders (last 10 by date), pending refunds, saved addresses", "ReactApiController.buildCustomerContext() / ChatController", "Controller+Service", "Yes"),
    (175, "Local Fallback Reply (No API Key)", "700+ lines of keyword-based responses; role-specific cross-question redirects; data extraction from context block strings", "AiAssistantService.localReply()", "Service", "Yes"),
    (176, "AI Chat Page (Web)", "Thymeleaf template chat-widget.html; session-based; ChatController resolves role (customer/vendor/admin/guest) and builds context", "ChatController", "Controller", "Yes"),
    (177, "AI Chat API (React/Flutter)", "POST /api/react/assistant/chat — JWT-auth; X-Customer-Id optional (guest fallback); returns reply, role, name", "ReactApiController", "Controller", "Yes"),
]

rows += [
    ("GEOCODING & LOCATION", "", "", "", ""),
    (178, "IP-based Auto PIN Detection", "GET /api/geocode/auto — ip-api.com lookup; extracts countryCode, city, zip; validates Indian PIN; falls back to postalpincode.in by city name; private IPs rejected", "GeocodingController", "Controller", "Yes"),
    (179, "Coordinate-to-PIN (GPS)", "GET /api/geocode/pin?lat=&lon= — India bounds check; Nominatim (OSM) → BigDataCloud → postalpincode.in fallback cascade; 3-stage with error isolation", "GeocodingController", "Controller", "Yes"),
    (180, "City-to-PIN Lookup", "GET /api/geocode/by-city?city= — postalpincode.in API; returns first valid Pincode", "GeocodingController", "Controller", "Yes"),
    (181, "Server-side HTTP Proxy", "All geo APIs called server-side; avoids CORS; no API key exposure; retry across multiple APIs transparently; 4s connect + 5s read timeouts", "GeocodingController.httpGet()", "Controller", "Yes"),
    (182, "JSON Parser (no Jackson)", "extractJson(body, key) — manual regex for both string and numeric values; firstNonBlank() utility", "GeocodingController", "Controller", "Yes"),
]

rows += [
    ("ADMIN — ACCOUNTS & USER MANAGEMENT", "", "", "", ""),
    (183, "Admin Account Dashboard (Web+API)", "Get all accounts with metadata (totalOrders, totalSpent); search by name/email; stats (total/active/suspended)", "AdminAccountService / AdminAccountController", "Service+Controller", "Yes"),
    (184, "Toggle Account Active Status", "AdminAccountService.toggleAccountStatus() — activate/deactivate customer; logs to stdout", "AdminAccountService", "Service", "Yes"),
    (185, "Delete Account (Admin)", "AdminAccountService.deleteAccount() — deletes wishlist→refunds→orders→customer cascade; also via /api/react/admin/accounts/{id}", "AdminAccountService / ReactApiController", "Service+Controller", "Yes"),
    (186, "Admin Generate Password Reset Link", "OTP generated, email sent (sendPasswordResetByAdmin); returns resetUrl + otp", "AdminAccountService", "Service", "Yes"),
    (187, "Admin Change User Role (Web)", "GET /admin/security — RBAC page; POST role change; updates session if currently logged-in user", "UserAdminService / EkartController", "Service+Controller", "Yes"),
    (188, "Admin Change User Role (API)", "PATCH /api/react/admin/users/{id}/role + /api/flutter/admin/users/{id}/role — validates Role enum", "ReactApiController / UserAdminApiController", "Controller", "Yes"),
    (189, "Admin Search Users (API)", "GET /api/react/admin/users/search?q=&type= — searches customers+vendors+delivery boys; type filter", "ReactApiController", "Controller", "Yes"),
    (190, "Admin Toggle Customer/Vendor Active (API)", "POST /api/flutter/admin/customers/{id}/toggle-active + /api/react/admin/customers/{id}/toggle-active", "FlutterApiController / ReactApiController", "Controller", "Yes"),
    (191, "Admin View User Profile", "GET /api/flutter/admin/accounts/{id}/profile — recentOrders (10), wishlistItems, totalSpent, avgOrderValue", "AdminAccountService / FlutterApiController", "Service+Controller", "Yes"),
]

rows += [
    ("MISCELLANEOUS & INFRASTRUCTURE", "", "", "", ""),
    (192, "CloudinaryHelper", "saveToCloudinary() (product images, profile images), saveBannerToCloudinary() (1200×375, crop fill, auto quality), saveVideoToCloudinary() (video resource_type)", "CloudinaryHelper (helper)", "Helper", "Yes"),
    (193, "EmailSender (Async, Thymeleaf templates)", "20+ email methods: OTP, orderConfirmation, stockAlert, replacement, cancellation, backInStock, passwordReset, deliveryBoyOtp, deliveryOtp (doorstep), shippedEmail, deliveryConfirmation, deliveryBoyApproved/Rejected, warehouseChangeApproved/Rejected, disputeNotification", "EmailSender (helper)", "Helper", "Yes"),
    (194, "MessageRemover", "Spring @Component; removes success/failure session attributes after display; used in Thymeleaf templates", "MessageRemover (helper)", "Helper", "Yes"),
    (195, "GlobalExceptionHandler", "404 → 404.html; all exceptions: JSON for AJAX/non-GET/API, HTML for browser page loads; detects via Accept header, X-Requested-With, method, path prefix", "GlobalExceptionHandler (controller)", "Controller", "Yes"),
    (196, "ErrorService (ControllerAdvice)", "Handles NoResourceFoundException→404.html, MissingRequestHeaderException→400 JSON, all Exception→error.html", "ErrorService (service)", "Service", "Yes"),
    (197, "DbHealthController", "GET /api/health — counts customers + products; returns JSON status", "DbHealthController", "Controller", "Yes"),
    (198, "GenericReadOnlyController", "GET /api/{entity} + /api/{entity}/{id} — auto-resolves JpaRepository by name; for Flutter/React dev debugging", "GenericReadOnlyController", "Controller", "Yes"),
    (199, "UserActivityController", "POST /api/user-activity/batch — logs List<Map> activity entries (userId, actionType, metadata, timestamp); GET /api/user-activity/user/{userId} — last 20", "UserActivityController", "Controller", "Yes"),
    (200, "TestDataController", "GET /admin/test-data/load — creates demo vendors, products, customers, orders for testing", "TestDataController", "Controller", "Yes"),
    (201, "DotenvConfig", "Loads .env file into system properties at startup; makes ADMIN_EMAIL, AES_SECRET, CLOUDINARY_* etc. available to @Value", "DotenvConfig (config)", "Config", "Yes"),
    (202, "WebMvcConfig — Interceptor Registration", "Registers IndiaOnlyInterceptor (order 1), AuthGuard (order 2), DeliveryBoyAuthGuard (order 3); excludePathPatterns for static assets", "WebMvcConfig (config)", "Config", "Yes"),
    (203, "application.properties Configuration Keys", "spring.datasource.*, spring.mail.*, cloudinary.*, razorpay.key.id, admin.email/password, aes.secret/salt, jwt.secret, gemini.api.key, reporting.datasource.*, ekart.banner.*", "application.properties", "Config", "Yes"),
    (204, "Guest Browse Mode", "GuestService: session attr guest=true; loadGuestBrowse(), guestSearch(); no cart/wishlist; redirects to register", "GuestService / EkartController", "Service+Controller", "Yes"),
    (205, "Spring Boot Application Entry", "@SpringBootApplication; @EnableAsync; @EnableCaching — enables @Cacheable, @Async annotations", "EkartApplication", "Application", "Yes"),
    (206, "Admin Password Change via API", "POST /api/react/admin/change-password — validates currentPassword, updates in-memory adminPassword field AND writes to .env file (ADMIN_PASSWORD=REDACTED
    (207, "Vendor Code Auto-generation", "Generated as VND-NNNNN (5 digits) after first save using vendor.getId()", "VendorService", "Service", "Yes"),
    (208, "Delivery Boy Code Auto-generation", "DB-NNNNN (5 digits) after first save; warehouseCode WH-NNN after warehouse save", "DeliveryBoyService / DeliveryAdminService", "Service", "Yes"),
    (209, "FlutterAdminApiController", "Separate controller for admin-specific Flutter endpoints: delivery management, warehouse transfers, bulk product approval, pending dashboard stats", "FlutterAdminApiController", "Controller", "Yes"),
    (210, "Admin Refund Management (Web Page)", "AdminRefundController: list pending/all refunds; approve (REFUNDED status) / reject with reason; AdminReviewController: review list + delete", "AdminRefundController / AdminReviewController", "Controller", "Yes"),
    (211, "Admin Policy Management Page", "AdminPolicyPageController + PolicyPageController; React-based SPA for policy CRUD at /policies", "AdminPolicyPageController / PolicyPageController", "Controller", "Yes"),
]

# Write rows
rnum = 3
for row in rows:
    if isinstance(row[0], str):
        # Category separator
        ws1.merge_cells(f"B{rnum}:F{rnum}")
        ws1.cell(rnum, 1, "").fill = FILLS["cat"]; ws1.cell(rnum,1).border = border()
        c = ws1.cell(rnum, 2, f"  {row[0]}")
        c.fill = FILLS["cat"]; c.font = YELLOW; c.alignment = left; c.border = border()
        ws1.row_dimensions[rnum].height = 18
    else:
        nr, cat, feat, cls, layer, working = row
        w = working == "Yes"
        drow(ws1, rnum, [nr, cat, feat, cls, layer, working], w)
        ws1.row_dimensions[rnum].height = 30
    rnum += 1

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: HTML / Frontend Features
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("HTML Frontend Features")
ws2.sheet_properties.tabColor = "1b5e20"
ws2.freeze_panes = "A3"

set_col_widths(ws2, [5, 28, 55, 22, 12])

ws2.merge_cells("A1:E1")
c = ws2.cell(1, 1, "EKART — HTML/JS/CSS Frontend Feature Inventory (Thymeleaf + Static Assets)")
c.fill = FILLS["h1"]; c.font = WHITE; c.alignment = center
ws2.row_dimensions[1].height = 28

hdrs2 = ["#", "Category", "Feature / UI Element / Behaviour", "File(s)", "Working?"]
hrow(ws2, 2, hdrs2,
     [FILLS["h2b"]]*5,
     [WHITE]*5,
     [center]*5, 22)

rows2 = []
rows2 += [
    ("LAYOUT & NAVIGATION", "", "", ""),
    (1, "Common Navbar (Thymeleaf Fragment)", "Logo, search bar, cart icon with badge, wishlist, profile dropdown, login/logout; role-aware links (admin/vendor/customer)", "customer-home.html, customer-view-products.html (th:replace fragments)", "Yes"),
    (2, "common-layout.css", "Sticky navbar, responsive layout helpers, card shadows, custom scrollbar, ekart brand colors (#f5a623 amber)", "static/css/common-layout.css", "Yes"),
    (3, "ekart-ui.css", "Primary CSS: gradient backgrounds, product card hover animations, star rating display, badge styles, tab navigation, skeleton loader placeholders", "static/css/ekart-ui.css", "Yes"),
    (4, "Responsive Design", "Bootstrap 5 grid throughout all pages; mobile-first breakpoints; hamburger nav on mobile", "All HTML templates", "Yes"),
    (5, "Dark Mode Hint (CSS vars)", "CSS custom properties used for color tokens; UI supports theming via variable override", "ekart-ui.css", "Yes"),
    (6, "403 Error Page", "Custom 403 Forbidden page with navigation back to appropriate role home", "403.html", "Yes"),
    (7, "404 Error Page", "Minimal 404 page; auto-returned by GlobalExceptionHandler for NoResourceFoundException", "404.html", "Yes"),
    (8, "Generic Error Page", "error.html shows errorMessage attribute from GlobalExceptionHandler", "error.html", "Yes"),
    (9, "Blocked Page (India-only)", "Displayed for non-Indian IPs; explains geographic restriction; contact info", "blocked.html", "Yes"),
]

rows2 += [
    ("HOME & LANDING PAGES", "", "", ""),
    (10, "Public Landing Page (home.html)", "Pre-login page; banner carousel; category grid; product preview; guest browse link; vendor/customer login buttons", "home.html", "Yes"),
    (11, "Customer Home Page (post-login)", "Banner carousel (Cloudinary images); filterable product grid by parent/sub-category; cart count badge; recently viewed bar; product cards with stock/discount badges", "customer-home.html", "Yes"),
    (12, "Banner Carousel", "Auto-sliding with manual controls; Cloudinary-hosted images; link overlay; show on Home vs showOnCustomerHome toggle", "home.html, customer-home.html", "Yes"),
    (13, "Category Filter Panel (Customer Home)", "Parent categories with emojis; sub-category pills; AJAX product filtering without page reload", "customer-home.html", "Yes"),
    (14, "Product Card", "Image (Cloudinary URL), name, price with ₹, MRP strikethrough + discount %, stock badge (Out of Stock overlay), average star rating, wishlist heart icon, Add to Cart / View Details buttons", "customer-home.html, customer-view-products.html, search.html", "Yes"),
    (15, "Guest Browse Page", "Approved products without login requirement; 'Login to Add to Cart' CTA; search functionality", "guest-browse.html", "Yes"),
]

rows2 += [
    ("AUTHENTICATION PAGES", "", "", ""),
    (16, "Customer Register Page", "Name, email, mobile, password, confirm password fields; client-side validation; OTP flow redirect", "customer-register.html", "Yes"),
    (17, "Customer Login Page", "Email/password; OAuth2 social login buttons (Google/GitHub/Facebook/Instagram); forgot password link; links to register", "customer-login.html", "Yes"),
    (18, "Customer OTP Verification", "6-box OTP input (auto-focus-next, paste-to-fill); countdown timer; resend link", "customer-otp.html", "Yes"),
    (19, "Customer Forgot Password Page", "Email input; server-side OTP trigger; redirect to reset page", "customer-forgot-password.html", "Yes"),
    (20, "Customer Reset Password Page", "OTP + new password + confirm; regex rules displayed; success redirect to login", "customer-reset-password.html", "Yes"),
    (21, "Vendor Register / Login / OTP / Forgot / Reset", "Mirror of customer auth flow; vendor-specific branding", "vendor-register.html, vendor-login.html, vendor-otp.html, vendor-forgot-password.html, vendor-reset-password.html", "Yes"),
    (22, "Admin Login Page", "Email/password only; no OAuth; session-based; redirects to admin home on success", "admin-login.html", "Yes"),
    (23, "Delivery Boy Register Page", "Name, email, mobile, password, confirm, warehouse dropdown (AJAX-loaded active warehouses); OTP email flow", "delivery-register.html", "Yes"),
    (24, "Delivery Boy Login Page", "Email/password; handles all states: unverified→OTP resend, inactive→error, pending→pending page, success→home", "delivery-login.html", "Yes"),
    (25, "Delivery Boy OTP Page", "6-box OTP input; verifies email; if adminApproved=false → pending page; if approved → login", "delivery-otp.html", "Yes"),
    (26, "Delivery Boy Pending Page", "Waiting for admin approval message; instructions; no action required from user", "delivery-pending.html", "Yes"),
]

rows2 += [
    ("PRODUCT PAGES", "", "", ""),
    (27, "Product Detail Page", "Multi-image carousel (main + extras); video player (if videoLink set); price/MRP/discount; stock status; GST breakdown; Add to Cart / Notify Me (back-in-stock); vendor info; star rating bar; review list with images; similar products section", "product-detail.html", "Yes"),
    (28, "Star Rating Display", "5-star visual bar with filled/half/empty stars; review count; per-star breakdown bars (5★→1★)", "product-detail.html, customer-view-products.html", "Yes"),
    (29, "Back-in-Stock Subscribe Button", "Shown only when product.stock==0; 'Notify Me' button → AJAX POST to /notify-me/{id}; toggles to 'Unsubscribe' on subscription", "product-detail.html", "Yes"),
    (30, "Customer View Products (Grid)", "All approved products; filter by category; product cards with all attributes", "customer-view-products.html", "Yes"),
    (31, "Search Page", "Search results from AJAX /api/search query; spell-correction banner ('Did you mean X?'); no-results state", "search.html", "Yes"),
    (32, "Add Product Page (Vendor)", "Form: name, description, price, MRP, category, stock, stock alert threshold, pin code restriction, main image, 5 extra images, video; Cloudinary direct upload", "add-product.html", "Yes"),
    (33, "Edit Product Page (Vendor)", "Pre-filled form; optional image/video re-upload; pin code field; save → re-approval not required (existing flow)", "edit-product.html", "Yes"),
    (34, "Admin View Products Page", "All products (approved+pending); approve/reject toggle; vendor info; filter tabs", "admin-view-products.html", "Yes"),
    (35, "Vendor View Products (Manage)", "Vendor's products list; edit/delete actions; stock badge; approval status", "vendor-view-products.html", "Yes"),
    (36, "Vendor Storefront Page", "Public store info: vendor name, code, product count, stock alert badge; profile update form (name, mobile)", "vendor-store-front.html", "Yes"),
]

rows2 += [
    ("CART & CHECKOUT PAGES", "", "", ""),
    (37, "View Cart Page", "Item list with image, name, price, AJAX qty +/- buttons, remove; real-time total update; delivery charge indicator (Free if ≥₹500); 'Proceed to Pay' button; recommended products row", "view-cart.html", "Yes"),
    (38, "AJAX Cart Quantity Controls", "ekart-form-spinner.js — increments/decrements via /api/cart/increase|decrease; updates lineTotal, cartTotal, deliveryCharge, freeDelivery flag without page reload; debounced", "static/js/ekart-form-spinner.js + view-cart.html", "Yes"),
    (39, "Payment Page", "Cart summary; Razorpay checkout integration (opens Razorpay modal); COD option; delivery time selector (Standard/Express); delivery address selector with saved addresses; PIN code input with GPS auto-detect; GST breakdown (taxableBase + GST label); recommended products", "payment.html", "Yes"),
    (40, "GPS / IP Auto-detect PIN", "JS navigator.geolocation.getCurrentPosition() → /api/geocode/pin; fallback /api/geocode/auto (IP-based); auto-fills pin input + city label; manual override available", "payment.html (inline JS)", "Yes"),
    (41, "Order Success Page", "Order ID(s) for multi-vendor split; total amount; GST breakdown; delivery time; payment mode; confetti animation; track order CTA; delivery time estimate", "order-success.html", "Yes"),
    (42, "Coupon Input (Cart/Payment)", "Coupon code text input + Apply button; success shows discount amount + description; remove button; real-time total recalculation", "view-cart.html, payment.html", "Yes"),
]

rows2 += [
    ("ORDER & TRACKING PAGES", "", "", ""),
    (43, "View Orders Page", "Order list with tracking status badge; replacementRequested indicator; cancel button (if not delivered/cancelled); request replacement button (7-day window); reorder button; refund status display", "view-orders.html", "Yes"),
    (44, "Order History Page", "Chronological order list; amount, date, status; compact view for reference", "order-history.html", "Yes"),
    (45, "Track Orders Page", "All orders with progress bar (0-100% from TrackingStatus.getProgressPercent()); step indicators (Processing→Packed→Shipped→Out for Delivery→Delivered); real event-based history", "track-orders.html", "Yes"),
    (46, "Track Single Order Page", "Detailed timeline for one order; full TrackingEventLog history with timestamps and cities; delivery boy info; OTP delivery confirmation badge", "track-single-order.html", "Yes"),
    (47, "Vendor Orders Page", "Split into Pending (PROCESSING) / In Progress (PACKED/SHIPPED/OUT_FOR_DELIVERY) / Delivered; Mark as Packed AJAX button per order", "vendor-orders.html", "Yes"),
    (48, "Admin Delivery Management Page", "PACKED orders ready for assignment; eligible delivery boy dropdown per order (PIN-filtered AJAX); SHIPPED orders list; OUT_FOR_DELIVERY list; pending approval badges", "admin-delivery-management.html", "Yes"),
    (49, "Reorder Button (View Orders)", "Opens check-stock modal: per-item availability + current prices; confirm → AJAX POST reorder; shows out-of-stock items; cart redirect on success", "view-orders.html (inline JS)", "Yes"),
]

rows2 += [
    ("PROFILE & ACCOUNT PAGES", "", "", ""),
    (50, "Customer Profile Page", "Name, email, mobile display; edit form; profile image with upload/remove; OAuth provider badge (if linked); link/unlink social account buttons; recent activity", "customer-proflie.html", "Yes"),
    (51, "Address Management Page", "List saved addresses (structured: name, house/street, city, state, PIN); add new address form; delete button per address; PIN validation hint", "address-page.html", "Yes"),
    (52, "Customer Security Settings", "Change password form (current + new + confirm); 2FA info placeholder", "customer-security-settings.html", "Yes"),
    (53, "Customer Settings Page", "Links to profile, address, security, spending, wishlist, stock alerts; account deletion with confirmation modal", "CustomerSettingsController → views", "Yes"),
    (54, "Wishlist Page", "Grid of wishlisted products; stock status; remove button; add-to-cart quick action; empty state", "wishlist.html", "Yes"),
    (55, "Stock Alerts Page (Vendor)", "List all alerts sorted by unacknowledged first; alert message with product name, stock level, threshold; Acknowledge button; unacknowledged count badge", "stock-alerts.html", "Yes"),
    (56, "Spending Analytics Page (Customer)", "Monthly bar chart (Chart.js); category pie/bar chart; totalSpent, totalOrders, avgOrderValue, topCategory KPI cards", "spending.html", "Yes"),
    (57, "User Spending Page (Admin)", "Per-user spending breakdown; sortable table; categorySpending details", "user-spending.html", "Yes"),
    (58, "My Spending Page", "Variant spending page for customer", "my-spending.html", "Yes"),
]

rows2 += [
    ("ADMIN PAGES", "", "", ""),
    (59, "Admin Home Page", "Navigation dashboard: products, users, delivery, content, analytics, refunds, reviews, coupons, security, policies", "admin-home.html", "Yes"),
    (60, "Admin Products Page", "All products (approved+unapproved); approve/reject toggle; vendor info column; search/filter", "admin-view-products.html", "Yes"),
    (61, "Admin Content Page (Banners)", "Upload/add banners; toggle active/showOnHome/showOnCustomerHome; delete; displayOrder; preview image", "admin-content.html", "Yes"),
    (62, "Admin Accounts Page", "Customer account list; toggle active/suspend; delete; role change; search; profile deep-link", "admin-accounts.html", "Yes"),
    (63, "Admin User Search Page", "Live search (JS-powered) across customers + vendors by name/email; filter by type", "admin-user-search.html", "Yes"),
    (64, "Admin Coupon Management Page", "List all coupons; add new coupon form (code, type, value, minOrderAmount, maxDiscount, usageLimit, expiryDate); toggle active; delete", "admin-coupons.html", "Yes"),
    (65, "Admin Delivery Management Page", "PACKED/SHIPPED/OUT_FOR_DELIVERY order lists; assign delivery boy dropdown; approve/reject pending boys; warehouse management", "admin-delivery-management.html", "Yes"),
    (66, "Admin Warehouse Page", "Warehouse CRUD; name, city, state, PIN codes (comma-separated); active toggle; delivery boys assigned count", "admin-warehouse.html", "Yes"),
    (67, "Admin Refund Management Page", "Pending refunds list; approve/reject with reason; order details; amount; customer info", "admin-refunds.html", "Yes"),
    (68, "Admin Review Management Page", "All reviews; star rating filter; search by product/customer/comment; individual delete; bulk delete by product", "admin-review-managment.html", "Yes"),
    (69, "Admin Security/RBAC Page", "User list with role badges; role change dropdowns; customer/orderManager/admin counts; provider column", "admin-security.html", "Yes"),
    (70, "Admin Policies Page", "React SPA embedded for policy CRUD (create/edit/delete policies by slug/category/title)", "admin-policies.html", "Yes"),
    (71, "Analytics Page (Admin)", "Platform stats: customer/vendor/product/order counts; revenue; orders by status; daily order chart (7-day Chart.js bar); category distribution chart", "analytics.html", "Yes"),
]

rows2 += [
    ("DELIVERY BOY PAGES", "", "", ""),
    (72, "Delivery Home Page", "Profile card (name, code, warehouse, assigned PINs); To Pick Up (SHIPPED) cards; Out Now (OUT_FOR_DELIVERY) cards with OTP input; Delivered history; warehouse change request form", "delivery-home.html", "Yes"),
    (73, "Delivery OTP Input", "6-digit input on delivery-home for OUT_FOR_DELIVERY orders; AJAX POST to confirm delivery; shows success/error toast", "delivery-home.html (inline JS)", "Yes"),
    (74, "Warehouse Change Request UI", "Dropdown of active warehouses; reason textarea; submit button; shows pending badge if request active", "delivery-home.html", "Yes"),
]

rows2 += [
    ("EMAIL TEMPLATES", "", "", ""),
    (75, "OTP Email Template", "HTML email: Ekart branding, OTP code in highlighted box, expiry note", "otp-email.html (template)", "Yes"),
    (76, "Order Confirmation Email", "Order #ID, amount, payment mode, delivery time, item list table", "order-email.html (template)", "Yes"),
    (77, "Order Cancellation Email", "Order #ID, amount, items cancelled, refund timeline note", "cancel-email.html (template)", "Yes"),
    (78, "Replacement Request Email", "Order #ID, items, return instructions", "replacement-email.html (template)", "Yes"),
    (79, "Shipped Email", "Order #ID, delivery boy name, current city, items", "shipped-email.html (template)", "Yes"),
    (80, "Delivered Confirmation Email", "Order #ID, delivered message, review CTA", "delivered-email.html (template)", "Yes"),
    (81, "Delivery OTP Email (Doorstep)", "6-digit OTP highlighted box, order #ID, 'Show this to your delivery boy' instruction", "delivery-otp-email.html (template)", "Yes"),
    (82, "Back-in-Stock Email", "Product image + name + price + stock count; Buy Now button link", "back-in-stock-email.html (template)", "Yes"),
    (83, "Stock Alert Email (Vendor)", "Product name, current stock, threshold, product ID link", "stock-alert-email.html (template)", "Yes"),
    (84, "Delivery Boy Pending Approval Email (Admin)", "New delivery boy details card; link to /admin/delivery", "EmailSender (inline HTML)", "Yes"),
    (85, "Delivery Boy Approved/Rejected Email", "Approval: code + warehouse; Rejection: reason box", "EmailSender (inline HTML)", "Yes"),
    (86, "Warehouse Change Approved/Rejected Email", "New warehouse name+city (approved) or reason box (rejected); admin note", "EmailSender (inline HTML)", "Yes"),
    (87, "Order Dispute Admin Notification", "Order #ID, customer email, reason, description, timestamp; sent to admin email", "EmailSender (inline HTML)", "Yes"),
    (88, "Customer Refund Report Page", "HTML page showing customer's all refund requests with status", "customer-refund-report.html", "Yes"),
]

rows2 += [
    ("JAVASCRIPT FEATURES (Static)", "", "", ""),
    (89, "Toast Notification System (ekart-toast.js)", "createToast(type, title, message, duration); types: success/error/warning/info; slide-in/out animations; auto-dismiss; stack management; global window.EkartToast API", "static/js/ekart-toast.js", "Yes"),
    (90, "Skeleton Loading (ekart-skeleton.js)", "showSkeletons(container, count, type) / hideSkeletons(); types: card, list, table; CSS animated shimmer placeholders; used while AJAX data loads", "static/js/ekart-skeleton.js", "Yes"),
    (91, "Form Spinner (ekart-form-spinner.js)", "Cart qty +/- AJAX handlers; debounced requests; updates lineTotal, cartTotal, deliveryCharge, freeDelivery in DOM; spinner overlay on button during request", "static/js/ekart-form-spinner.js", "Yes"),
    (92, "Recently Viewed Products (recently-viewed.js)", "localStorage-based product ID list (max 10); deduplication; server sync on login (POST /api/recently-viewed/sync); load from server on page load (GET /api/recently-viewed/sync); renders product bar from /api/recently-viewed/products?ids=; horizontal scroll", "static/js/recently-viewed.js", "Yes"),
    (93, "User Action Logger (user-action-logger.js)", "Batch logs user actions (PAGE_VIEW, PRODUCT_VIEW, ADD_TO_CART, SEARCH, etc.) to /api/user-activity/batch; 5-second flush interval; userId from session; LocalDateTime timestamps", "static/js/user-action-logger.js", "Yes"),
    (94, "Search Suggestions (inline JS)", "Debounced input on search bar; GET /api/search/suggestions; renders dropdown with productName, category, imageLink; keyboard navigation (↑↓ Enter); auto-submit on selection", "search.html / customer-home.html (inline JS)", "Yes"),
    (95, "Wishlist Heart Toggle (inline JS)", "Heart icon AJAX POST /api/wishlist/toggle; optimistic UI update; toast notification; updates wishlist count badge", "customer-home.html, product-detail.html (inline JS)", "Yes"),
    (96, "Add to Cart AJAX (inline JS)", "POST /api/flutter/cart/add (Flutter) or session-based /add-to-cart/{id}; stock validation; cart badge update; toast notification", "customer-home.html, customer-view-products.html (inline JS)", "Yes"),
    (97, "AI Chat Widget Integration", "Floating chat button; opens chat widget iframe/overlay; passes session context; role-aware greeting", "ai-assistant-widget.html, chat-widget.html", "Yes"),
    (98, "Chat Widget UI", "Message bubbles (user/assistant); typing indicator; history preserved in session; role-aware intro; markdown-like bold rendering", "chat-widget.html", "Yes"),
    (99, "Razorpay Modal Integration (inline JS)", "Loads Razorpay SDK; creates order on backend; opens modal; handles payment.success/failure; submits form with razorpay_payment_id, razorpay_order_id", "payment.html (inline JS)", "Yes"),
    (100, "GPS + IP Auto-detect PIN (inline JS)", "navigator.geolocation.getCurrentPosition() → /api/geocode/pin?lat=&lon=; fallback /api/geocode/auto; fills PIN input + shows city label; manual override if auto-detect fails", "payment.html (inline JS)", "Yes"),
    (101, "Delivery Boy AJAX Actions (inline JS)", "Mark Pickup → POST /delivery/order/{id}/pickup; OTP Confirm → POST /delivery/order/{id}/confirm; Warehouse Change Request → POST /delivery/warehouse-change/request; all return JSON toasts", "delivery-home.html (inline JS)", "Yes"),
    (102, "Chart.js Integration (Vendor Sales)", "Bar chart for daily/weekly/monthly revenue; dynamic dataset from ordersJson server-rendered; responsive; color-coded bars", "vendor-sales-report.html (inline JS)", "Yes"),
    (103, "Chart.js Integration (Admin Analytics)", "Daily orders bar chart; category distribution chart; revenue breakdown", "analytics.html (inline JS)", "Yes"),
    (104, "Chart.js Integration (Customer Spending)", "Monthly spending bar chart; category spending pie/doughnut chart; KPI number animations", "spending.html (inline JS)", "Yes"),
    (105, "Product Image Gallery (inline JS)", "Main image swapper on thumbnail click; zoom on hover; extra images carousel", "product-detail.html (inline JS)", "Yes"),
    (106, "Product Video Player", "Native HTML5 video player; controls; shown only if videoLink is set", "product-detail.html", "Yes"),
    (107, "Coupon Code Apply/Remove (inline JS)", "POST /api/react/cart/coupon; updates discount display, cart total, deliveryCharge; DELETE removes coupon", "view-cart.html, payment.html (inline JS)", "Yes"),
    (108, "Reorder Modal (inline JS)", "GET /api/orders/{id}/check-stock; renders per-item availability table with status badges; Cancel / Confirm Reorder; POST /api/orders/{id}/reorder on confirm; redirect to cart", "view-orders.html (inline JS)", "Yes"),
    (109, "Refund Request Modal (inline JS)", "Type select (Refund/Replacement); reason textarea; file upload (evidence images JPEG/PNG/WEBP max 5); POST refund request then upload images", "view-orders.html (inline JS)", "Yes"),
    (110, "Report Issue Modal (inline JS)", "Reason text + description; POST /api/react/orders/{id}/report-issue; success toast", "view-orders.html (inline JS)", "Yes"),
    (111, "Admin Delivery Assignment Dropdown (inline JS)", "On order click GET /admin/delivery/boys/for-order/{id}; renders eligible delivery boys (PIN-filtered) in dropdown; submit assigns", "admin-delivery-management.html (inline JS)", "Yes"),
]

rows2 += [
    ("RESPONSIVE / UX FEATURES", "", "", ""),
    (112, "Success/Failure Flash Messages", "Thymeleaf th:if on session 'success'/'failure' attributes; auto-cleared by MessageRemover; styled banners", "All Thymeleaf pages", "Yes"),
    (113, "Pagination (where applicable)", "Order lists, product grids have scroll-load or pagination for large datasets", "view-orders.html, admin pages", "Yes"),
    (114, "Cart Item Count Badge (Navbar)", "Integer badge on cart icon; server-rendered from customer.cart.items.size(); updated live via AJAX", "customer-home.html + all customer pages", "Yes"),
    (115, "Wishlist Count Badge", "Count shown in navbar; updated on toggle via AJAX response", "customer-home.html", "Yes"),
    (116, "Stock Alert Count Badge (Vendor Navbar)", "Unacknowledged alerts count shown in vendor navbar", "vendor-home.html", "Yes"),
    (117, "Discount Badge on Product Card", "Shows '20% OFF' badge when mrp > price; calculated server-side via Product.getDiscountPercent()", "customer-home.html, product cards", "Yes"),
    (118, "Out of Stock Overlay", "Grayed overlay on product card image; 'Out of Stock' label; Add to Cart disabled", "All product grid pages", "Yes"),
    (119, "Delivery Charge Progress Indicator", "Shows 'Add ₹X more for free delivery' when cart < ₹500; disappears at ₹500+", "view-cart.html", "Yes"),
    (120, "Print Invoice / Receipt", "Order detail pages have print-friendly CSS; browser print dialog", "view-orders.html, order-success.html", "Yes"),
    (121, "Mobile Hamburger Menu", "Bootstrap 5 navbar-toggler collapses nav on mobile; smooth transition", "All pages", "Yes"),
    (122, "Dark Overlay / Modal Backdrop", "Custom modal backdrop for confirmation dialogs; reorder, refund, dispute modals", "view-orders.html, payment.html", "Yes"),
    (123, "Empty State Designs", "Illustrated empty states for cart, wishlist, orders, search no-results, stock alerts", "Various pages", "Yes"),
    (124, "Sample Product CSV (Admin/Vendor)", "sample-product-upload.csv — downloadable template for bulk product upload", "static/sample-product-upload.csv", "Yes"),
]

rnum2 = 3
for row in rows2:
    if isinstance(row[0], str):
        ws2.merge_cells(f"B{rnum2}:E{rnum2}")
        ws2.cell(rnum2, 1, "").fill = FILLS["cat"]; ws2.cell(rnum2,1).border = border()
        c = ws2.cell(rnum2, 2, f"  {row[0]}")
        c.fill = FILLS["cat"]; c.font = YELLOW; c.alignment = left; c.border = border()
        ws2.row_dimensions[rnum2].height = 18
    else:
        nr, cat, feat, f, working = row
        w = working == "Yes"
        drow(ws2, rnum2, [nr, cat, feat, f, working], w)
        ws2.row_dimensions[rnum2].height = 30
    rnum2 += 1

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: Summary Dashboard
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Summary Dashboard")
ws3.sheet_properties.tabColor = "4a148c"
set_col_widths(ws3, [30, 20, 20, 30])

ws3.merge_cells("A1:D1")
c = ws3.cell(1,1,"EKART — Feature Summary Dashboard")
c.fill = FILLS["h1"]; c.font = WHITE; c.alignment = center
ws3.row_dimensions[1].height = 30

summary_data = [
    ("JAVA BACKEND", ""),
    ("Auth & Security", 26),
    ("Product Management", 22),
    ("Cart & Checkout", 21),
    ("Order Management", 20),
    ("Delivery System", 14),
    ("Refund System", 8),
    ("Reviews", 9),
    ("Wishlist", 6),
    ("Back-in-Stock Notifications", 6),
    ("Stock Alerts", 6),
    ("Profile & Address", 12),
    ("Banners & Content", 8),
    ("Analytics & Reporting", 13),
    ("AI Assistant", 6),
    ("Geocoding & Location", 5),
    ("Admin — Accounts & User Management", 9),
    ("Miscellaneous & Infrastructure", 20),
    ("TOTAL JAVA FEATURES", 211),
    ("", ""),
    ("HTML/JS/CSS FRONTEND", ""),
    ("Layout & Navigation", 9),
    ("Home & Landing Pages", 6),
    ("Authentication Pages", 11),
    ("Product Pages", 10),
    ("Cart & Checkout Pages", 7),
    ("Order & Tracking Pages", 7),
    ("Profile & Account Pages", 9),
    ("Admin Pages", 13),
    ("Delivery Boy Pages", 3),
    ("Email Templates", 14),
    ("JavaScript Features (Static)", 23),
    ("Responsive / UX Features", 13),
    ("TOTAL FRONTEND FEATURES", 125),
    ("", ""),
    ("GRAND TOTAL ALL FEATURES", 336),
]

r = 2
for label, count in summary_data:
    is_total = "TOTAL" in str(label) or "GRAND" in str(label)
    is_section = label in ("JAVA BACKEND", "HTML/JS/CSS FRONTEND")
    is_blank = label == ""
    if is_blank:
        r += 1
        continue
    c1 = ws3.cell(r, 1, label)
    c2 = ws3.cell(r, 2, count if count else "")
    if is_section:
        c1.fill = FILLS["h2"]; c1.font = WHITE; c2.fill = FILLS["h2"]; c2.font = WHITE
    elif is_total:
        c1.fill = FILLS["h1"]; c1.font = WHITE
        c2.fill = FILLS["h1"]; c2.font = WHITE
    else:
        c1.fill = FILLS["row_a"] if r%2==0 else FILLS["row_b"]
        c2.fill = FILLS["yes"]
        c2.font = bold_font(11)
    c1.alignment = left; c2.alignment = center
    c1.border = border(); c2.border = border()
    ws3.row_dimensions[r].height = 20
    r += 1
wb.save("EKART_Full_Feature_Sheet.xlsx")
print("Done — workbook saved")
